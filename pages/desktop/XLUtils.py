import openpyxl
from openpyxl.styles import PatternFill, Font
import xlwings as xw


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    # sheet = workbook.get_sheet_by_name(sheetName)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName, r):
    workbook = openpyxl.load_workbook(file)
    # sheet = workbook.get_sheet_by_name(sheetName)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    # sheet = workbook.get_sheet_by_name(sheetName)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    # wb = xw.Book(file)
    # workbook = openpyxl.load_workbook(file)
    # # sheet = workbook.get_sheet_by_name(sheetName)
    # sheet = workbook[sheetName]
    # cell = sheet.cell(row=rownum, column=columnno)
    # cell.value = data
    # if data.strip() == "Test Passed":
    #     cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    # elif data.strip() == "Test failed":
    #     cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    # workbook.save(file)
    # wb.save(file)

    # approach 2
    wb = xw.Book(file)
    sheet = wb.sheets[sheetName]
    cell = sheet.range((rownum, columnno))  # Specify the cell by row and column number
    cell.value = data

    # Set cell color based on the data value
    if data.strip() == "Test Passed":
        cell.color = (0, 255, 0)  # Green
    elif data.strip() == "Test failed":
        cell.color = (255, 0, 0)  # Red
    wb.save(file)  # Save changes to the file


def column_letter_to_index(letter):
    """Convert Excel column letter(s) to a 1-based column index."""
    index = 0
    for char in letter:
        index = index * 26 + (ord(char.upper()) - ord('A') + 1)
    return index


# utils/logger_utils.py
import logging
import os

def setup_logger(name="main_logger"):
    # Create logs directory if it doesn't exist
    log_dir = "D:\\logs"
    os.makedirs(log_dir, exist_ok=True)

    # Define log file path
    log_file = os.path.join(log_dir, "login_tests.log")

    # Create logger
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)

    # Prevent duplicate logs
    if not logger.handlers:
        # Create file handler
        file_handler = logging.FileHandler(log_file)
        file_handler.setLevel(logging.INFO)

        # Create console handler
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)

        # Define formatter (same format for both handlers)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)

        # Add handlers to logger
        logger.addHandler(file_handler)
        logger.addHandler(console_handler)

    return logger
